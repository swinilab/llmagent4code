"""One-off converter: docs/OMS_TestCases_BVA_EP_EN.xlsx -> docs/OMS_TestCases_BVA_EP_EN.md"""
import openpyxl

SRC = "docs/OMS_TestCases_BVA_EP_EN.xlsx"
DST = "docs/OMS_TestCases_BVA_EP_EN.md"


def esc(v):
    if v is None:
        return ""
    s = str(v)
    return s.replace("|", "\\|").replace("\n", "<br>")


def sheet_to_table(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""
    header = rows[0]
    lines = []
    lines.append("| " + " | ".join(esc(h) for h in header) + " |")
    lines.append("|" + "|".join(["---"] * len(header)) + "|")
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        lines.append("| " + " | ".join(esc(c) for c in row) + " |")
    return "\n".join(lines)


def readme_to_md(ws):
    lines = []
    for row in ws.iter_rows(values_only=True):
        a, b = (row + (None, None))[:2]
        if a is None and b is None:
            lines.append("")
        elif b is None:
            lines.append(f"{a}")
        else:
            lines.append(f"- **{a}**: {b}")
    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out = ["# OMS Test Case Suite (BVA + EP)", ""]

    out.append("## README")
    out.append("")
    out.append(readme_to_md(wb["README"]))
    out.append("")

    out.append("## Summary")
    out.append("")
    out.append(sheet_to_table(wb["Summary"]))
    out.append("")

    for name in ["Customer", "Product", "Order", "Payment", "Invoice"]:
        out.append(f"## {name}")
        out.append("")
        out.append(sheet_to_table(wb[name]))
        out.append("")

    with open(DST, "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print(f"wrote {DST}")


if __name__ == "__main__":
    main()
